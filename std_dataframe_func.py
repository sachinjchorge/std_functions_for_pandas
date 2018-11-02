# -*- coding: utf-8 -*-
"""
Created on Tue Sep 13 14:04:21 2016

@author: chorge
"""

import pandas as pd
import datetime
import numpy as np

# data visualization
import seaborn as sns
from matplotlib import pyplot as plt
plt.rcParams['figure.figsize']= (10,4)
from matplotlib import style

#Alogrithms
from sklearn import linear_model
from sklearn.linear_model import LogisticRegression
from sklearn.ensemble import RandomForestClassifier
from sklearn.linear_model import Perceptron
from sklearn.linear_model import SGDClassifier
from sklearn.tree import DecisionTreeClassifier
from sklearn.neighbors import KNeighborsClassifier
from sklearn.svm import SVC, LinearSVC
from sklearn.naive_bayes import GaussianNB
#from sklearn.preprocessing import LabelEncoder

import colorama
from colorama import Fore, Back, Style


# Basic Functions

def get_new_List_Item(X_Item):
    Next_X_Column = input("Get Next X_Column (Update X to Exit) : ")
    if Next_X_Column != "X":
        X_Item.append(Next_X_Column)
        print(X_Item)
        get_new_List_Item(X_Item)
    
    else:
        
        print(X_Item)
        return X_Item
    return X_Item

def get_only_numeric_part(db, ColumnName, NewColumnName):
    db[NewColumnName]= db[ColumnName].convert_objects(convert_numeric= True)
    return db
    
def get_only_string_part(db, ColumnName, NewColumnName):
    db[NewColumnName]= db[ColumnName].str.replace(r'[\(\)\d]+', '')
    return db

def convert_str_to_int(str_input):
    int_output = int(str_input)
    return int_output

def convert_num_to_str(num_input):
    str_output = str(num_input)
    return str_output

def convert_str_to_float(str_input):
    float_output = float(str_input)
    return float_output

def convert_pd_columns_str_to_float(df, colName):
    df[colName]= df[colName].convert_objects(convert_numeric= True)
    return df

def reversed_Loop(LoopName):
    Reversed_Loop = reversed(LoopName)
    return Reversed_Loop

def Loop_over_indices(LoopName):
    for i, color in enumerate(LoopName):
        print(i, '-->', color)

def Loop_over_2_Lists(L1,L2):
    for i,j in zip(L1,L2):
        print(i,'-->',j)

#    for i,j in izip(L2,L2):
#        print(i,'-->',j)
        
    for i in sorted(L1):
        print(i)
        
    for i in sorted(L2, reverse = True):
        print(i)

    for i in sorted(L1,key = len):
        print(i)
        
def find(seq, target):
    for i, value in enumerate(seq):
        if value== target:
            break
    else:
        return -1
#        print(i)
    return i


def convert_char_date_to_date(char_date):
    from datetime import datetime
    date_obj = datetime.strptime(char_date,'%b%d%Y%I:%M%p')
    return date_obj

def convert_float_to_str(df, columnName):
    df[columnName]= df[columnName].map(lambda x: '{0:.0f}'.format(x))
    df[columnName]= df[columnName].map(str)
    return df

# Functions related to extraction of Data
def get_filename():
    filename = input("Update filename for Data :")
    filename = filename.replace("/","//")
    return filename

def extractdf_from_file(filename, *sheetname):
    starttime = datetime.datetime.now()
    print(starttime)
    print('Capturing data frame')
    if '.csv' in filename:
        try:
            df = pd.read_csv(filename,encoding= 'ISO-8859-1', na_values= '',low_memory= False, parse_dates= True)
        except:
            df = pd.read_csv(filename,na_values= '')
        finally:
            print('Dataframe captured in Time : ', datetime.datetime.now()- starttime )
            return df
    elif '.xlsx' in filename:
        try:
            df= pd.read_excel(filename, sheetname,na_values= '' )
        except:
            df = pd.read_excel(filename, na_values= '')
        finally:
            print('Dataframe captured in Time : ', datetime.datetime.now()- starttime )
            return df
    elif '.html' in filename:
        try:
            df = pd.read_html(filename)
            print('Dataframe captured in Time : ', datetime.datetime.now()- starttime )
            return df
        except:
            print('Could not create dataframe')            
            pass
    elif '.pickle' in filename:
        try:
            df= Reload_PickleData(filename)
            print('Dataframe captured in Time : ', datetime.datetime.now()- starttime )
            return df
        except:
            print('Could not create dataframe')
            pass

def extractdf_from_SQL(SQLConnectionString, SQL_Select_Statement):
    import pyodbc as dbc
    starttime = datetime.datetime.now()
    print(starttime)
    print('Capturing data frame')
    try:
        connection = dbc.connect(SQLConnectionString)
        df = pd.read_sql(SQL_Select_Statement,connection)
        print('Dataframe captured in Time : ', datetime.datetime.now()- starttime )
        return df
    except:
        print('Could not create dataframe')            
        pass

        
def replace_blank_with_nan(df):
    df = df.apply(lambda x: x.str.strip() if isinstance(x, str) else x).replace('', np.nan)
    return df

def convert_to_str(df,ColNameList):
    for ColName in ColNameList:
        try:
            df[ColName]= df[ColName].astype(str)
        except:
            print("Error in converting to string for Column :", ColName)
    return df

def remove_data_from_dfCol(RemoveList, ColName, df):
    for lstit in RemoveList:
        try:
            df = df[df[ColName]!=lstit]
        except:
            pass
    return df

def convert_to_datetime_from_float_dfColName(df, ColNameList):
    date_time_convert = lambda x: datetime.datetime.utcfromtimestamp((float(x) - 25569) * 86400.0) if float(x) > 25569 else np.nan
    for ColName in ColNameList:
        try:
            df[ColName]= df[ColName].map(date_time_convert)
        except:
            pass
    return df

def extractymd_from_datetime_dfColName(df, ParentColName, NewColName, convType):
    if convType =="Year":
        conv = lambda x: x.year
    elif convType== "Month":
        conv = lambda x: x.month
    elif convType== "Day":
        conv = lambda x: x.day
    elif convType== "WeekDay":
        conv = lambda x: x.weekday()
    try:
        df[NewColName]= df[ParentColName].map(conv)
    except:
        pass
    return df

def get_timediff_for_dfCol(df, StartColName, EndColName, NewColName, DiffUnits ):
    # for Weeks, Days, Hours, Minutes, Seconds
    if DiffUnits== 'Weeks':
        try:
            df[NewColName]= (df[EndColName]- df[StartColName])/(datetime.timedelta(weeks=1))
        except:
            pass
    elif DiffUnits== 'Days':
        try:
            df[NewColName]= (df[EndColName]- df[StartColName])/(datetime.timedelta(days=1))
        except:
            pass
    elif DiffUnits=='Hours': 
        try:
            df[NewColName]= (df[EndColName]- df[StartColName])/(datetime.timedelta(hours=1))
        except:
            pass
    elif DiffUnits =='Minutes':
        try:
            df[NewColName]= (df[EndColName]- df[StartColName])/(datetime.timedelta(minutes=1))
        except:
            pass
    elif DiffUnits =='Seconds':
        try:
            df[NewColName]= (df[EndColName]- df[StartColName])/(datetime.timedelta(seconds=1))
        except:
            pass
    return df
    
def get_Bus_Days_for_dfCol(df, StartColName, EndColName, NewColName):
    try:
        df[NewColName]= np.busday_count(df[StartColName].values.astype('datetime64[D]'),df[EndColName].values.astype('datetime64[D]'))
    except:
        x1 = [x.date() for x in df[StartColName]]
        x2 = [x.strftime('%Y-%m-%d') for x in df[EndColName].fillna(0)]
        df['BusinessDays'] = np.busday_count(x1,x2)
        Daysstatus = lambda x: x if x >= 0.0 else np.nan
        df[NewColName]= df['BusinessDays'].map(Daysstatus)
    return df
    
def drop_columns_from_dfCol(df, ColName):
    try:
        df.drop(ColName,axis = 1, inplace = True )
    except:
        pass
    return df

def df_dtypes(df):
    print(df.dtypes)

def change_column_value_string(df, conditional_col, condition_string, new_col, value):
    try:
        df.loc[df[conditional_col]==condition_string, new_col]= value
    except:
        pass
    return df

def change_column_value_bycolcompare(df, first_col, condition_string, last_col,new_col, value):
    try:
        if condition_string == 'G':
            df.loc[df[last_col] > df[first_col], new_col]= value
        elif condition_string == 'GE':
            df.loc[df[last_col] >= df[first_col], new_col]= value
        elif condition_string == 'L':
            df.loc[df[last_col] < df[first_col], new_col]= value
        elif condition_string == 'LE':
            df.loc[df[last_col] <= df[first_col], new_col]= value
        else:
            df.loc[df[last_col] == df[first_col], new_col]= value
    except:
        pass
    return df

def change_column_value_byvalcompare(df, first_col, condition_string, cond_value ,new_col, value):
    try:
        if condition_string == 'G':
            df.loc[df[first_col] > cond_value, new_col]= value
        elif condition_string == 'GE':
            df.loc[df[first_col] >= cond_value, new_col]= value
        elif condition_string == 'L':
            df.loc[df[first_col] < cond_value, new_col]= value
        elif condition_string == 'LE':
            df.loc[df[first_col] <= cond_value, new_col]= value
        else:
            df.loc[df[first_col] == cond_value, new_col]= value
    except:
        pass
    return df

def findSpikes(data, threshold=0.2):
    prev = None
    for i, v in enumerate(data):
        if prev is None:
            prev = v
            continue

        delta = (v + prev) / 2
        if delta >= threshold:
            print("Found spike at index %d (value %f)" % (i, v))

        prev = v
        
def Save_PickleData(data, filename):
    import pickle
    with open(filename, 'wb') as handle:
        pickle.dump(data, handle, protocol = pickle.HIGHEST_PROTOCOL)
        print(".............Data Saved to Pickles :", filename)
    

def Reload_PickleData(filename):
    import pickle
    data = pickle.load(open(filename,'rb'))
    print("................Data uploaded from pickles :", filename)
    return data

def Drop_Unique_Val_Columns(df):
    clean_consts = [col for col in df.columns if df[col].nunique()==1]
    df.drop(clean_consts, axis=1 , inplace= True)
    print("This action will drop : ", len(clean_consts), " columns")
    print("All droppend columns : \n", clean_consts)
    return df

def get_null_details(df):
    total = df.isnull().sum().sort_values(ascending= False)
    percent_1= df.isnull().sum()/df.isnull().count()*100
    percent_2= round(percent_1,1).sort_values(ascending= False)
    missing_data = pd.concat([total,percent_2], axis= 1, keys = ['Total', '%']) 
    return missing_data

def convert_data_to_train_test_from_df(df, X_column, Y_column, testsize= 0.2, random_state= 0):
    from sklearn.model_selection import train_test_split

    if Y_column != df.columns[-1]:
        df= df.reindex(columns = (list([a for a in df.columns if a!= Y_column]) + [Y_column]))
    X= df[X_column].iloc[:].values
    Y = df.iloc[:,len(df.columns)-1].values
    
    if testsize== float(0):
        X_Train = X
        X_Test= X
        Y_Train = Y
        Y_Test = Y
    else:
        X_Train, X_Test, Y_Train, Y_Test = train_test_split(X, Y, test_size= testsize, random_state= 0)
    return X_Train, X_Test, Y_Train, Y_Test

def convert_data_to_train_test_from_Array(X, Y, testsize= 0.2, random_state= 0):
    from sklearn.model_selection import train_test_split
    if testsize== float(0):
        X_Train = X
        X_Test= X
        Y_Train = Y
        Y_Test = Y
    else:
        X_Train, X_Test, Y_Train, Y_Test = train_test_split(X, Y, test_size= testsize, random_state= 0)
    return X_Train, X_Test, Y_Train, Y_Test

def describe_numeric_and_nonnumeric(df):
    df_describe = df.describe(include= ['O'])
    return df_describe

"""
            ActivityName ApplicationName Core_NonCore LineItem1 LineItem2  \
count                708             708          708       708       708   
unique                10               3            3        44         6   
top     Query resolution            Idle         Core                       
freq                 232             604          505       213       703   

       LineItem3     Process ProcessName  UserName Masking ID Weekday Total  \
count        708         708         708       708        708     708   708   
unique         1           1           3         9          9       5     1   
top               BGL AP CRC              peyyalah        BA8     Tue  IDLE   
freq         708         708         604       124        124     150   708  

"""

def describe_df(df):
    df_describe= df.describe()
    return df_describe

"""
       CalculatedTime       Unit1   UnitValue       Unit2  UnitValue2  Unit3  \
count      708.000000  708.000000  708.000000  708.000000       708.0  708.0   
mean      1529.159605    4.538136    0.048023    1.024011         0.0    1.0   
std       1483.942869    5.001269    0.570822    0.306903         0.0    0.0   
min        180.000000    1.000000    0.000000    1.000000         0.0    1.0   
25%        434.500000    1.000000    0.000000    1.000000         0.0    1.0   
50%        995.500000    2.000000    0.000000    1.000000         0.0    1.0   
75%       2075.000000    6.000000    0.000000    1.000000         0.0    1.0   
max      10100.000000   45.000000   12.000000    6.000000         0.0    1.0   

"""

def Encode_Labels(df, columnName):
    from sklearn.preprocessing import LabelEncoder
    lbl= LabelEncoder()
    
    df[:,columnName]= lbl.fit_transform(df[:,columnName])
    return df

def Encode_Labels_One_Hot_Encoder(df, columnName):
    from sklearn.preprocessing import OneHotEncoder
    onehotencoder = OneHotEncoder(categorical_features= [columnName])
    df= onehotencoder.fit_transform(df).toarray()
    return df
    
def get_bins(df, columnName, binsize):
    n, bins, patches = plt.hist(df[columnName], bins=binsize, label='hst')
    Bins = pd.Series(bins)
    return Bins

def Convert_to_int(df, columnName):
    df[columnName]= df[columnName].astype(int)
    return df

def return_FacetGrid(df, columnName):
    return sns.FacetGrid(df, col= columnName)

def get_row_and_column_count(df):
    print(Fore.RED + "==============================================================")
    print(Fore.GREEN +'The Data has {0} rows and {1} columns'.format(df.shape[0], df.shape[1]))
    print(Fore.RED + "==============================================================")

def print_Head_and_Tail_of_Data(df):
    print_to_console_initial("Initial 5 Rows of Data for Understanding")
    print(df.head(5))
    print_to_console_end("Initial 5 Rows of Data for Understanding")
    
    print_to_console_initial("Last 5 Rows of Data for Understanding")
    print(df.tail(5))
    print_to_console_end("Last 5 Rows of Data for Understanding")

def print_df_info(df):
    print_to_console_initial("Basic Information of Captured DataFrame")
    print(df.info())
    print_to_console_end("Basic Information of Captured DataFrame")
    

#    sns.plt.show
    
def Plot_Distribution(df, columnName):
    print(Fore.RED + "==============================================================")
    print(Fore.RED + "Printing Distribution Plot for Output Variable :", columnName)
    sns.distplot(df[columnName])
    plt.show()
    
def check_Skewness(df,columnName):
    skew_value= df[columnName].skew()
    print("The skewness of ", columnName, " is {}".format(skew_value))
    return skew_value
    
def transform_to_normal_data(df, columnName):
    target = np.log(df[columnName])
    print('Skewness is :', target.skew())
    sns.distplot(target)
    plt.show()
    return target, target.skew()

def sep_data_into_num_and_cat(df):
    numeric_data = df.select_dtypes(include = [np.number])
    cat_data= df.select_dtypes(exclude= [np.datetime64, np.number] )
    dt_data= df.select_dtypes(include = [np.datetime64])
    print("There are {} numeric and {} categorical columns in data".format(numeric_data.shape[1], cat_data.shape[1]))
    return numeric_data, cat_data, dt_data

def plot_heatmap(numeric_data, target_col):
    corr = numeric_data.corr()
    sns.heatmap(corr)
    plt.savefig('Heatmap.png')
    plt.show()
    print(corr[target_col].sort_values(ascending= False))
    return corr

def anova(frame, target_col, cat):
    from scipy import stats
    anv = pd.DataFrame()
    anv['features']= cat
    pvals= []
    for c in cat:
        samples = []
        for cls in frame[c].unique():
            s= frame[frame[c]== cls][target_col].values
            samples.append(s)
        pval = stats.f_oneway(*samples)[1]
        pvals.append(pval)
    anv['pval']= pvals
    return anv.sort_values('pval')

def draw_anova(df, target_col):
    cat_data= df.select_dtypes(exclude = [np.number])
    cat = [f for f in df.columns if df.dtypes[f]=='object']
    cat_data[target_col]= df[target_col].values
    k= anova(cat_data,target_col,cat)
    k['disparity']= np.log(1./k['pval'].values)
    sns.barplot(data= k, x= 'features', y= 'disparity')
    plt.xticks(rotation= 90)
    plt.show()
  
def create_numeric_plots(df):
    num = [f for f in df.columns if df.dtypes[f] != 'object']
    nd= pd.melt(df, value_vars= num)
    n1 = sns.FacetGrid(nd, col= 'variable', col_wrap= 4, sharex= False, sharey= False )
    n1 = n1.map(sns.distplot, 'value')
    n1
    
def create_jointplt(df, x_colname, y_colname):
    sns.jointplot(x= df[x_colname], y = df[y_colname])
    plt.show()
    
def create_pivot_and_plot(df,index_column, target_column):
    pivot= df.pivot_table(index= index_column, values= target_column, aggfunc= np.sum)
#    pivot = pivot.sort_values(by= target_column)
    pivot.plot(kind= 'bar', color= 'red')
    plt.show()

def boxplot(df,x,y,exclude,**kwargs):
    try:
        if exclude== 0:
            sns.boxplot(x= df[x], y= df[y])
            x= plt.xticks(rotation= 90)
            plt.show()
        else:
            try:
                sns.boxplot(x= df[df[x]!= exclude][x], y= df[y])
                x= plt.xticks(rotation= 90)
                plt.show()
            except:
                print("Not able to print Box Plot, Pls check the variables and/or Exclude parameters")
    except:
        print("Not able to print Box Plot, Pls check the variables and/or Exclude parameters")
    
def plot_boxplot(df, target_column):
    cat= [f for f in df.columns if df.dtypes[f]=='object']
    p= pd.melt(df, id_vars= target_column, value_vars= cat)
    g= sns.FacetGrid(p, col= 'variable', col_wrap= 2, sharex= False, sharey= False, size= 5)
    try:
        g= g.map(boxplot, 'value', target_column)
        g
        plt.show()
    except ValueError:
        print("Skip plotting Box plots")
        
def plot_dynamic_pivot_and_plot(df, OutputVariable, cat_data):
    print(Fore.RED + "==========================================================================")
    print(Fore.RED + "Plot Bar Graphs for Output Variable against Categorical Variables")
    print(Fore.RED + "Create Plot against ", OutputVariable, " and plot for the DataFrame")
    print("Select from the following variable names ", cat_data.columns)
    # Create Bar Graphs for SUM
    print("Getting Ready to print Bar Graphs")
    Index_Variable = "XYZ"
    while Index_Variable!= "X":
        print("Select from the following variable names ", cat_data.columns)
        Index_Variable= input("Which column to plot against OutputVariable / Update X for Exiting the loop : ")
        if Index_Variable == "X":
            print("Done with Plotting for Bar Graphs on Variables")
            return
        else:
            try:
                create_pivot_and_plot(df, Index_Variable,OutputVariable)
            except KeyError:
                print("Wrong Variable Name Entered")
                print("Select from the following variable names ", cat_data.columns)


def plot_dynamic_boxplot(df, OutputVariable, cat_data):
    print(Fore.RED + "==========================================================================")
    print(Fore.RED + "Print Box Plots for Output Variable against Categorical Variables")
    print(Fore.RED + "Create Plot against ", OutputVariable, " and plot for the DataFrame")
    print("Select from the following variable names ", cat_data.columns)
    # Create Bar Graphs for SUM
    print("Getting Ready to print Box Plot Graphs")
    Index_Variable = "XYZ"
    while Index_Variable!= "X":
        print("Select from the following variable names ", cat_data.columns)
        Index_Variable= input("Which column to plot against OutputVariable / Update X for Exiting the loop : ")
        if Index_Variable == "X":
            print("Done with Plotting for Box Plot Graphs on Variables")
        else:
            try:
                exclude = 0
                boxplot(df, Index_Variable,OutputVariable, exclude)
                Response = input("Do you want to exclude any data from Categorical variable and recheck the Box Pot (Y/N) :")
                if Response == 'Y':
                    while exclude != 'X':
                        exclude = input("Update Categorical Data you want to exclude in the Box Plot / Update X to Exit Loop :")
                        try:
                            boxplot(df, Index_Variable,OutputVariable, exclude)
                        except:
                            exclude = input("Update Categorical Data you want to exclude in the Box Plot / Update X to Exit Loop :")
                
            except KeyError:
                print("Wrong Variable Name Entered")
                print("Select from the following variable names ", cat_data.columns)


def print_to_console_initial(text_data):
    print(Fore.RED + Style.NORMAL)
    print(Fore.RED + Style.BRIGHT +  "==========================================================================" + color.END)
    print(Fore.RED + Style.BRIGHT +"Initiating " + text_data)
    print(Fore.RED + "==========================================================================")
    print(Fore.BLACK + Style.NORMAL)


def print_to_console_end(text_data):
    print(Fore.RED + Style.NORMAL)
    print(Fore.RED + "==========================================================================")
    print(Fore.RED + Style.BRIGHT+ "Completed "+ text_data)
    print(Fore.RED + "==========================================================================")
    print(Fore.BLACK + Style.NORMAL)

class color:
   PURPLE = '\033[95m'
   CYAN = '\033[96m'
   DARKCYAN = '\033[36m'
   BLUE = '\033[94m'
   GREEN = '\033[92m'
   YELLOW = '\033[93m'
   RED = '\033[91m'
   BOLD = '\033[1m'
   UNDERLINE = '\033[4m'
   END = '\033[0m'

def merge_data_and_keep_common_and_uncommon_rows(df1,df2, list_CommonColumn):
    result1= pd.merge(df1,df2, how = 'outer', on= list_CommonColumn)
    result1= pd.concat([result1, pd.DataFrame(data= df1, columns= result1.columns) ])
    result1= result1.drop_duplicates()
    return result1

def create_empty_dataFrames(list_of_dfNames):
    x= [pd.DataFrame for x in list_of_dfNames]
    return x

def check_missing_values(df):
    print_to_console_initial("Missing Values in the DataFrame")
    print(df.columns[df.isnull().any()])
    miss = df.isnull().sum()#/len(df)
    miss= miss[miss>0]
    if len(miss)==0:
        print(Fore.GREEN + "No Missing Values in DataFrame")
        print_to_console_end("Missing Values in the DataFrame")
        return
    miss.sort_values(inplace= True)
    print(miss)
    
    
    #visualising missing values
    miss= miss.to_frame()
    miss.columns= ['count']
    miss.index.names= ['Name']
    miss['Name']= miss.index
    
#    #plot the missing value count
#    sns.set(style= "whitegrid", color_codes= True)
#    sns.barplot(x= 'Name', y= 'count', data = miss)
#    plt.xticks(rotation = 90)
#    plt.show()
    print_to_console_end("Missing Values in the DataFrame")
    return miss

def replace_missing_values(df):
    miss= check_missing_values(df)
        
    try:
        if len(miss)>0:
            for name in miss['Name']:
                print("Replace Empty Value in Column : ", name, " With data type : ", df[name].dtype)
                replace_na_value = input("Replace missing value with (0 / Mean / Mode / Median) : ")
                if df[name].dtype== 'object':
                    replace_na_value= replace_na_value
                elif df[name].dtype== '<M8[ns]':
                    df= df[df[name].notnull()]
                else:
                    if replace_na_value == '0':
                        replace_na_value = 0
                        replace_na_value= float(replace_na_value)
                    elif replace_na_value == 'Mean':
                        replace_na_value= df[name].mean()
                        replace_na_value= float(replace_na_value)
                    elif replace_na_value== 'Median':
                        replace_na_value = df[name].median()
                        replace_na_value= float(replace_na_value)
                    elif replace_na_value== 'Mode':
                        try:
                            replace_na_value = df[name].mode()
                            replace_na_value= float(replace_na_value)
                        except:
                            replace_na_value = df[name].median()
                            replace_na_value= float(replace_na_value)
                    else:
                        try:
                            replace_na_value= float(replace_na_value)
                        except:
                            pass
                try:        
                    df[name]= df[name].fillna(replace_na_value)
                except:
                    pass
    except TypeError:
        print(" No Missing values in Data")
        return df
                
    miss = check_missing_values(df)
    return df

def print_array_info(X):
    print_to_console_initial("Array info")
    if len(X)>= 10:
        print(X[1:10])
    else:
        print(X)
    print_to_console_end("Array info")

def OpenFilePath():
    import tkinter as tk
    from tkinter import filedialog
    
    root = tk.Tk()
    root.withdraw()
    
    file_path = filedialog.askopenfilename()
    return file_path

def mad_based_outlier(points, thresh=3.5):
    if len(points.shape) == 1:
        points = points[:,None]
    median = np.median(points, axis=0)
    diff = np.sum((points - median)**2, axis=-1)
    diff = np.sqrt(diff)
    med_abs_deviation = np.median(diff)

    modified_z_score = 0.6745 * diff / med_abs_deviation

    return modified_z_score > thresh

def percentile_based_outlier(data, threshold=95):
    diff = (100 - threshold) / 2.0
    minval, maxval = np.percentile(data, [diff, 100 - diff])
    return (data < minval) | (data > maxval)

def outliers_iqr(ys):
    quartile_1, quartile_3 = np.percentile(ys, [25, 75])
    iqr = quartile_3 - quartile_1
    lower_bound = quartile_1 - (iqr * 1.5)
    upper_bound = quartile_3 + (iqr * 1.5)
    return np.where((ys > upper_bound) | (ys < lower_bound))

def Create_Sub_Totals_from_MultiIndex_3_Levels(df1):
    df2 = df1.groupby(level=[0,1]).sum()
    df2.index = pd.MultiIndex.from_arrays([df2.index.get_level_values(0),
                                       df2.index.get_level_values(1) + ' Total',
                                       len(df2) * ['']])
    df3 = df1.groupby(level=[0]).sum()
    df3.index = pd.MultiIndex.from_arrays([df3.index.get_level_values(0) + ' Total',
                                       len(df3) * [''],
                                       len(df3) * ['']])
    df_final = pd.concat([df1,df2,df3]).sort_index()
    return df_final

def Create_Sub_Totals_from_MultiIndex_2_Levels(df1):
    df3 = df1.groupby(level=[0]).sum()
    df3.index = pd.MultiIndex.from_arrays([df3.index.get_level_values(0) + ' Total',
                                       len(df3) * [''],
                                       len(df3) * ['']])
    df_final = pd.concat([df1,df3]).sort_index()
    return df_final

def Create_Sub_Totals_from_Normal_df(df1, No_of_Levels):
    N= len(df1.columns)
    No_of_Levels= No_of_Levels-1
    X= N - No_of_Levels
    print(N, No_of_Levels, X)
    df_trial= pd.DataFrame()
    for i in range(0,No_of_Levels):
        if i == 0:
            by_List = [df1.columns[i]]
        else:
            by_List.append(df1.columns[i])
            print(by_List)
        for j in range(X, N):
            if j == X:
                df_trial = df1.groupby(by = by_List).agg({df1.columns[j]:'sum'})
            else:
                df_trial= pd.merge(df_trial,df1.groupby(by = by_List).agg({df1.columns[j]:'sum'}), on =by_List)
        if i<= No_of_Levels:
            for k in range(i+1, No_of_Levels+1):
                df_trial[df1.columns[k]]= ''
        df_temp= df_trial.reset_index()
        for l in range(0, N):
            if l==0:
                Column_List= [df1.columns[l]]
            else:
                Column_List.append(df1.columns[l])
        df_temp = df_temp[Column_List]
        if i ==0:
            df_Output = df_temp
        else:
            df_Output= df_Output.append(df_temp, ignore_index= True)
    df_Output = df_Output.append(df1)
    for i in range(1, No_of_Levels + 1):
        df_Output.loc[(df_Output[df_Output.columns[i]]=='') , df_Output.columns[i]]= 'Total'
    for i in range(1, No_of_Levels + 1):
        
        for k in range(i+1, No_of_Levels + 1):
            df_Output.loc[(df_Output[df_Output.columns[i]]=='Total') & (df_Output[df_Output.columns[k]]=='Total'), df_Output.columns[k]]= ''

    for l in range(0, No_of_Levels + 1):
        if l==0:
            Sort_List= [df1.columns[l]]
        else:
            Sort_List.append(df1.columns[l])
    df_Output= df_Output.sort_values(by = Sort_List )

def import_data_from_url(url):
    df= pd.read_csv(url, parse_dates= True, error_bad_lines= False)
    return df

def get_sheetnames_from_xlsx(ExcelPathName):
    import openpyxl
    wb = openpyxl.load_workbook(ExcelPathName)
    sheet_name_list = wb.sheetnames
    return sheet_name_list

def replace_null_values(df):
    df.loc[:, df.dtypes == object] = df.loc[:, df.dtypes == object].fillna('NA')
    df.loc[:, df.dtypes == '<M8[ns]'] = df.loc[:, df.dtypes == '<M8[ns]'].fillna(datetime.datetime.now())
    df.loc[:, df.dtypes == float] = df.loc[:, df.dtypes == float].fillna(0)
    return df

def apply_tariff_cut(df):
    cents_per_kwh = pd.cut(x=df.index.hour,
                           bins=[0, 7, 17, 24],
                           include_lowest=True,
                           labels=[12, 20, 28]).astype(int)
    df['cost_cents'] = cents_per_kwh * df['energy_kwh']

def apply_tariff_digitize(df):
    prices = np.array([12, 20, 28])
    bins = np.digitize(df.index.hour.values, bins=[7, 17, 24])
    df['cost_cents'] = prices[bins] * df['energy_kwh'].values    
    
def snap_rounding(x):
    import math
    EPSILON = 0.000001
    snap_ceil = lambda x: math.ceil(x) if abs(x - round(x)) > EPSILON else round(x)
    snap_floor = lambda x: math.floor(x) if abs(x - round(x)) > EPSILON else round(x)
    return snap_ceil, snap_floor

def extract_word_from_string():
    words = lambda text: ''.join(c if c.isalnum() else ' ' for c in text).split()
    return words

def print_in_place():
    import time
    import sys
    for progress in range(100):
      time.sleep(0.1)
      sys.stdout.write("Download progress: %d%%   \r" % (progress) ) 
      sys.stdout.flush()
      
def store_data_to_hd5(base_dir, df, dfname):
    # Create storage object with filename `processed_data`
    data_store = pd.HDFStore(base_dir + '\\PRS_APJ_Data.h5')
    
    # Put DataFrame into the object setting the key as 'preprocessed_df'
    data_store['preprocessed_df'] = df
    data_store.close()    

def access_data_from_hd5():
# Access data store
    data_store = pd.HDFStore('processed_data.h5')
    
    # Retrieve data using key
    preprocessed_df = data_store['preprocessed_df']
    data_store.close()
    
def access_data_tables_from_hd5():   
    data_store.keys()

def Remove_extra_spaces_from_col_name(db):
    print("..Renaming Columns to remove extra spaces")
    db.rename(columns = lambda x: x.strip(), inplace= True)
    db= db.rename(columns = lambda x: x.replace(".00",''))
    print("....Columns renamed by removing extra spaces")
    return db

def Remove_spaces_in_data(db):
    # Remove spaces and errors in text data
    print("..Cleaning Errors in Data")
    db.loc[:, db.dtypes == object] = db.loc[:, db.dtypes == object].replace(',','')
    for j in range(5):
        db.loc[:, db.dtypes == object] = db.loc[:, db.dtypes == object].replace('  ',' ')
    db.loc[:, db.dtypes == object] = db.loc[:, db.dtypes == object].replace('&','AND')
    db.loc[:, db.dtypes == object] = db.loc[:, db.dtypes == object].replace('\r+','')
    
    for i in db.columns:
        try:
            db[i]= db[i].map(lambda x: x.strip())
        except:
            pass
    del i
    print("....Completed Cleaning Errors in Data")
    return db

def read_files_from_folder(filepath, filetype):
    import glob
    counter = 0
    for file_name in glob.glob(filepath +"\\*" + ".csv"):
        print("Reading Data from file: ",file_name)
        df1= pd.read_csv(file_name, encoding = 'ISO-8859-1')
        print("Data frame loaded, size :", len(df1))
        df1.rename(columns = lambda x:x.strip(), inplace= True)
        df1= Remove_spaces_in_data(df1)
        if counter==0:
            df= df1
        else:
            df= pd.concat([df,df1], ignore_index= True)
        counter = counter + 1
    return df

def CalcOutlier(df, colName):
    df_Num = df[colName]
    df_mean, df_std = np.mean(df_Num), np.std(df_Num)
    cut = df_std * 3
    lower, upper = df_mean - cut, df_mean + cut
    outliers_lower = [x for x in df_Num if x<lower]
    outliers_higher = [x for x in df_Num if x> upper]
    outliers_total = [x for x in df_Num if x<lower or x > upper]
    outliers_removed = [x for x in df_Num if x>lower and x<upper]
    print('For column ', colName,' Identified lowest outliers: %d' % len(outliers_lower)) # printing total number of values in lower cut of outliers
    print('For column ', colName,' Identified upper outliers: %d' % len(outliers_higher)) # printing total number of values in higher cut of outliers
    print('For column ', colName,' Identified outliers: %d' % len(outliers_total)) # printing total number of values outliers of both sides
    print('For column ', colName,' Non-outlier observations: %d' % len(outliers_removed)) # printing total number of non outlier values
    print('For column ', colName,' Total percentual of Outliers: ', round((len(outliers_total) / len(outliers_removed) )*100, 4)) # Percentual of outliers in points
    
    return    

def Plot_Distribution_with_log(df,colName):
    # Printing some statistics of our data
    print(colName," Min Value: ", 
          df[df[colName] > 0][colName].min()) # printing the min value
    print(colName," Mean Value: ", 
          df[df[colName] > 0][colName].mean()) # mean value
    print(colName," Median Value: ", 
          df[df[colName] > 0][colName].median()) # median value
    print(colName," Max Value: ", 
          df[df[colName] > 0][colName].max()) # the max value
    
    plt.figure(figsize = (14,5))
    
    plt.subplot(1,2,1)
    ax = sns.distplot(np.log(df[df[colName] > 0][colName] + 0.01), bins=40, kde=True)
    ax.set_xlabel(colName +'Log', fontsize=15) #seting the xlabel and size of font
    ax.set_ylabel('Distribution', fontsize=15) #seting the ylabel and size of font
    ax.set_title("Distribution of " + colName, fontsize=20) #seting the title and size of font
    
    # setting the second plot of our grid of graphs
    plt.subplot(1,2,2)
    # ordering the total of users and seting the values of transactions to understanding 
    plt.scatter(range(df.shape[0]), np.sort(df[colName].values))
    plt.xlabel('Index', fontsize=15) # xlabel and size of words
    plt.ylabel(colName, fontsize=15) # ylabel and size of words
    plt.title(colName +" Distribution", fontsize=20) # Setting Title and fontsize
    
    plt.show()    
    return

def Get_First_Word_from_Col(df, colName, NewcolName):
    import re
    df[NewcolName]= df[colName].apply(lambda x: re.search(' ([A-Z][a-z]+)\.', x).group(1))
    return df

def Neural_Network_Model():
    from keras.models import Sequential
    from keras.layers import Dense, Activation, Dropout
    import keras
    from keras.optimizers import SGD
    import graphviz  
    
    # Creating the model
    model = Sequential()
    
    # Inputing the first layer with input dimensions
    model.add(Dense(100, 
                    activation='relu',  
                    input_dim=20,
                    kernel_initializer='uniform'))
    #The argument being passed to each Dense layer (18) is the number of hidden units of the layer. 
    # A hidden unit is a dimension in the representation space of the layer.
    
    #Stacks of Dense layers with relu activations can solve a wide range of problems
    #(including sentiment classification), and you’ll likely use them frequently.
    
    # Adding an Dropout layer to previne from overfitting
    model.add(Dropout(0.50))
    
    #adding second hidden layer 
    model.add(Dense(100,
                    kernel_initializer='uniform',
                    activation='relu'))
    
    # Adding another Dropout layer
    model.add(Dropout(0.50))
    
    # adding the output layer that is binary [0,1]
    model.add(Dense(1,
                    kernel_initializer='uniform',
                    activation='sigmoid'))
    #With such a scalar sigmoid output on a binary classification problem, the loss
    #function you should use is binary_crossentropy
    
    #Visualizing the model
    model.summary()   

    #Creating an Stochastic Gradient Descent
    sgd = SGD(lr = 0.01, momentum = 0.9)
    
    # Compiling our model
    model.compile(optimizer = sgd, 
                       loss = 'binary_crossentropy', 
                       metrics = ['accuracy'])
    #optimizers list
    #optimizers['SGD', 'RMSprop', 'Adagrad', 'Adadelta', 'Adam', 'Adamax', 'Nadam']
    
    # Fitting the ANN to the Training set
    model.fit(X_Train, Y_Train, batch_size= 60, epochs = 30, verbose = 2)
    scores = model.evaluate(X_Train, Y_Train, batch_size=30)
    print("%s: %.2f%%" % (model.metrics_names[1], scores[1]*100))
# Fit the model
    history = model.fit(X_Train, Y_Train, validation_split=0.20, 
                        epochs=180, batch_size=10, verbose=0)
    
    # list all data in history
    print(history.history.keys())
    # summarizing historical accuracy
    plt.plot(history.history['acc'])
    plt.plot(history.history['val_acc'])
    plt.title('Model Accuracy')
    plt.ylabel('Accuracy')
    plt.xlabel('Epoch')
    plt.legend(['train', 'test'], loc='upper left')
    plt.show()

def reduce_mem_usage(df):
    """ iterate through all the columns of a dataframe and modify the data type
        to reduce memory usage.        
    """
    start_mem = df.memory_usage().sum() / 1024**2
    print('Memory usage of dataframe is {:.2f} MB'.format(start_mem))

    for col in df.columns:
        col_type = df[col].dtype

        if col_type != object:
            c_min = df[col].min()
            c_max = df[col].max()
            if str(col_type)[:3] == 'int':
                if c_min > np.iinfo(np.int8).min and c_max < np.iinfo(np.int8).max:
                    df[col] = df[col].astype(np.int8)
                elif c_min > np.iinfo(np.int16).min and c_max < np.iinfo(np.int16).max:
                    df[col] = df[col].astype(np.int16)
                elif c_min > np.iinfo(np.int32).min and c_max < np.iinfo(np.int32).max:
                    df[col] = df[col].astype(np.int32)
                elif c_min > np.iinfo(np.int64).min and c_max < np.iinfo(np.int64).max:
                    df[col] = df[col].astype(np.int64)  
            else:
                if c_min > np.finfo(np.float16).min and c_max < np.finfo(np.float16).max:
                    df[col] = df[col].astype(np.float16)
                elif c_min > np.finfo(np.float32).min and c_max < np.finfo(np.float32).max:
                    df[col] = df[col].astype(np.float32)
                else:
                    df[col] = df[col].astype(np.float64)

    end_mem = df.memory_usage().sum() / 1024**2
    print('Memory usage after optimization is: {:.2f} MB'.format(end_mem))
    print('Decreased by {:.1f}%'.format(100 * (start_mem - end_mem) / start_mem))

    return df
